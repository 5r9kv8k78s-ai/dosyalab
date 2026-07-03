from pathlib import Path

import fitz
import pytest
from openpyxl import load_workbook

from app.modules.converter import get_converter
from app.modules.converter.pdf_to_xlsx import PdfToXlsxConverter


def test_pdf_to_xlsx_is_registered_automatically() -> None:
    assert isinstance(get_converter("pdf-to-xlsx"), PdfToXlsxConverter)


def test_convert_preserves_tables_across_multiple_pages(
    sample_pdf_path: Path, tmp_path: Path
) -> None:
    """The real fixture has tables on two separate pages (verified directly
    with PyMuPDF's find_tables(): page 3 has one 9x2 table, page 5 has a 6x2
    and an 8x4 table) — a genuine multi-page, multi-table-per-page case, not
    an assumption.
    """
    output_path = PdfToXlsxConverter().convert(sample_pdf_path, tmp_path)

    assert output_path.exists()
    assert output_path.suffix == ".xlsx"

    workbook = load_workbook(output_path)

    # One sheet per page that actually contains a table; pages without
    # tables (1, 2, 4, 6) are skipped rather than producing empty sheets.
    assert workbook.sheetnames == ["Page 3", "Page 5"]

    page3 = workbook["Page 3"]
    assert page3.max_column == 2
    assert [cell.value for cell in page3[1]] == ["73%", "64%"]

    # Page 5 has two tables (6x2 and 8x4) separated by a blank row: 6 rows +
    # 1 blank + 8 rows = 15 rows, and columns span the wider of the two (4).
    page5 = workbook["Page 5"]
    assert page5.max_row == 15
    assert page5.max_column == 4
    assert [cell.value for cell in page5[1]][:2] == ["Class", "Heuristics for classification"]
    assert all(cell.value is None for cell in page5[7])  # separator row between tables


def test_convert_raises_when_no_tables_found(sample_pdf_path: Path, tmp_path: Path) -> None:
    # A real single-page subset of the fixture (page 1, verified via
    # find_tables() to contain zero tables) rather than a synthetic PDF.
    source = fitz.open(sample_pdf_path)
    single_page = fitz.open()
    single_page.insert_pdf(source, from_page=0, to_page=0)
    no_tables_pdf = tmp_path / "no-tables.pdf"
    single_page.save(no_tables_pdf)
    single_page.close()
    source.close()

    with pytest.raises(ValueError, match="No tables"):
        PdfToXlsxConverter().convert(no_tables_pdf, tmp_path / "out")
