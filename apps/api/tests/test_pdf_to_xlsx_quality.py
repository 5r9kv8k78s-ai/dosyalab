"""V2-3 conversion quality tests for `PdfToXlsxConverter` (see
app/modules/converter/pdf_to_xlsx.py). Distinct from test_pdf_to_xlsx_
converter.py's structural regression tests (exact sheet names/dimensions/
known cell values) — these assert *observable output quality* properties
against the repository's one real PDF fixture with tables
(tests/fixtures/sample.pdf: a 9x2 table on page 3, a 6x2 and an 8x4 table
on page 5 — verified directly with PyMuPDF's find_tables(), not assumed).

No cell content is asserted here beyond what's already public in
test_pdf_to_xlsx_converter.py — only structural/quality properties.
"""

from pathlib import Path

from openpyxl import load_workbook

from app.modules.converter.base import VerificationResult
from app.modules.converter.pdf_to_xlsx import PdfToXlsxConverter


def _convert(sample_pdf_path: Path, tmp_path: Path):
    output_path = PdfToXlsxConverter().convert(sample_pdf_path, tmp_path)
    return output_path, load_workbook(output_path)


def test_workbook_opens_with_at_least_one_worksheet(sample_pdf_path: Path, tmp_path: Path) -> None:
    _, workbook = _convert(sample_pdf_path, tmp_path)
    assert len(workbook.sheetnames) >= 1


def test_no_worksheet_is_fully_empty(sample_pdf_path: Path, tmp_path: Path) -> None:
    _, workbook = _convert(sample_pdf_path, tmp_path)
    for name in workbook.sheetnames:
        sheet = workbook[name]
        non_empty = [
            cell.value for row in sheet.iter_rows() for cell in row if cell.value not in (None, "")
        ]
        assert non_empty, f"worksheet {name!r} has no non-empty cells"


def test_worksheet_names_are_unique_and_excel_safe(sample_pdf_path: Path, tmp_path: Path) -> None:
    _, workbook = _convert(sample_pdf_path, tmp_path)
    names = workbook.sheetnames
    assert len(names) == len(set(names)), "duplicate worksheet names"
    forbidden_chars = set(":\\/?*[]")
    for name in names:
        assert 1 <= len(name) <= 31, f"{name!r} violates Excel's 31-character sheet name limit"
        assert not (forbidden_chars & set(name)), f"{name!r} contains an Excel-unsafe character"


def test_multi_table_separator_row_stays_blank(sample_pdf_path: Path, tmp_path: Path) -> None:
    """Guards a design decision made in this phase: page 5's two tables are
    deliberately separated by one fully-blank row (see convert()'s
    row_cursor logic) — a generic "strip fully-empty rows" cleanup pass
    would silently destroy this multi-table boundary, so it was NOT
    implemented (see the V2-3 report's Aşama 6 section). This test would
    catch that regression if it were ever added."""
    _, workbook = _convert(sample_pdf_path, tmp_path)
    page5 = workbook["Page 5"]
    assert page5.max_row == 15
    assert all(cell.value is None for cell in page5[7])


def test_extracted_cell_text_has_no_raw_newlines_or_untrimmed_whitespace(
    sample_pdf_path: Path, tmp_path: Path
) -> None:
    """Page 5's first table has wrapped multi-line labels — PyMuPDF's
    find_tables().extract() returns some of them with an embedded newline
    (verified directly against the real fixture, not assumed). Every
    extracted string cell must come out whitespace-normalized: no raw
    "\\n"/"\\t", and no leading/trailing whitespace.
    """
    _, workbook = _convert(sample_pdf_path, tmp_path)
    checked_any = False
    for name in workbook.sheetnames:
        for row in workbook[name].iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value:
                    checked_any = True
                    assert "\n" not in cell.value, f"raw newline survived in {name!r}"
                    assert "\t" not in cell.value, f"raw tab survived in {name!r}"
                    assert cell.value == cell.value.strip(), f"untrimmed whitespace in {name!r}"
    assert checked_any, "fixture produced no string cells to check"


def test_successful_conversion_still_passes_v2_1_verify(
    sample_pdf_path: Path, tmp_path: Path
) -> None:
    output_path, _ = _convert(sample_pdf_path, tmp_path)
    result = PdfToXlsxConverter().verify(output_path)
    assert result == VerificationResult(ok=True, reason=None)
