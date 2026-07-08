"""`PdfToXlsxConverter.verify()` (see app/modules/converter/pdf_to_xlsx.py) —
V2-1's second real output verifier. Deliberately NOT a re-implementation of
`convert()`'s existing `tables_written == 0` business rule (that already
guarantees at least one non-empty worksheet before any file is ever saved —
an empty workbook can't be saved at all). This checks the actual bytes
written to disk instead: is it still a valid, openable XLSX container with
at least one worksheet.
"""

import re
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.modules.converter.base import VerificationResult
from app.modules.converter.pdf_to_xlsx import PdfToXlsxConverter


def _zero_worksheet_xlsx(tmp_path: Path) -> Path:
    """A valid ZIP/XLSX container with zero worksheets — openpyxl itself
    refuses to *save* one directly (`IndexError: At least one sheet must
    be visible`), so this is built by saving a normal one-sheet workbook
    and then independently stripping the <sheets> entries from its
    xl/workbook.xml, the on-disk shape a write that failed partway through
    could plausibly leave behind. Verified: openpyxl's read_only loader
    tolerates this and reports sheetnames == [] rather than raising.
    """
    one_sheet = Workbook()
    one_sheet.active.title = "OnlySheet"
    valid_output = tmp_path / "one-sheet.xlsx"
    one_sheet.save(valid_output)

    no_sheets_output = tmp_path / "no-sheets.xlsx"
    with zipfile.ZipFile(valid_output) as zin, zipfile.ZipFile(no_sheets_output, "w") as zout:
        for item in zin.namelist():
            data = zin.read(item)
            if item == "xl/workbook.xml":
                data = re.sub(rb"<sheets>.*?</sheets>", b"<sheets/>", data)
            zout.writestr(item, data)
    return no_sheets_output


def test_verify_ok_true_on_a_real_successful_conversion(
    sample_pdf_path: Path, tmp_path: Path
) -> None:
    output_path = PdfToXlsxConverter().convert(sample_pdf_path, tmp_path)

    result = PdfToXlsxConverter().verify(output_path)

    assert result == VerificationResult(ok=True, reason=None)


def test_verify_rejects_a_missing_output_file(tmp_path: Path) -> None:
    result = PdfToXlsxConverter().verify(tmp_path / "never-written.xlsx")

    assert result.ok is False
    assert result.reason == "output_missing"


def test_verify_rejects_bytes_that_are_not_a_valid_xlsx_container(tmp_path: Path) -> None:
    fake_output = tmp_path / "not-a-workbook.xlsx"
    fake_output.write_bytes(b"this is not a zip/xlsx container")

    result = PdfToXlsxConverter().verify(fake_output)

    assert result.ok is False
    assert result.reason == "invalid_xlsx"


def test_verify_rejects_a_workbook_with_zero_worksheets(tmp_path: Path) -> None:
    # convert() itself can never actually produce this on disk (the
    # tables_written == 0 guard raises before workbook.save() is reached) —
    # see _zero_worksheet_xlsx for how this on-disk shape is constructed.
    no_sheets_output = _zero_worksheet_xlsx(tmp_path)

    result = PdfToXlsxConverter().verify(no_sheets_output)

    assert result.ok is False
    assert result.reason == "no_worksheets"


def test_zero_worksheet_fixture_is_itself_a_valid_openable_container(tmp_path: Path) -> None:
    """Guards the test fixture above against silently rotting into a
    generic invalid_xlsx failure instead of exercising the no_worksheets
    branch specifically."""
    no_sheets_output = _zero_worksheet_xlsx(tmp_path)

    workbook = load_workbook(no_sheets_output, read_only=True)
    assert workbook.sheetnames == []
    workbook.close()
