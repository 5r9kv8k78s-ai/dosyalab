import logging
from pathlib import Path

import fitz
from openpyxl import Workbook, load_workbook

from app.modules.converter.base import ConversionModule, VerificationResult
from app.modules.converter.registry import register_converter

logger = logging.getLogger(__name__)


def _normalize_cell_value(value: object) -> object:
    """Collapses embedded newlines/tabs/repeated whitespace from a wrapped
    PDF label (e.g. `"X-axis\\nvalue"` -> `"X-axis value"`, verified against
    the real sample.pdf fixture's page 5 table) into a single space, and
    strips leading/trailing whitespace. Never touches a non-string value or
    an already-empty string — this only cleans up text within a cell, it
    never adds, removes, or reorders any cell.
    """
    if isinstance(value, str) and value.strip():
        return " ".join(value.split())
    return value


class PdfToXlsxConverter(ConversionModule):
    """Converts PDF documents to XLSX by extracting tables via PyMuPDF's
    built-in table finder (`Page.find_tables()`, available since PyMuPDF
    1.23) — no new heavy dependency, since PyMuPDF is already used for PDF
    validation elsewhere in the app.

    One worksheet per source page that contains at least one table; a page
    with multiple tables gets all of them on that page's sheet, separated by
    a blank row, preserving each table's row/column structure. Pages with no
    tables are skipped entirely. If the PDF has no tables anywhere,
    conversion fails with a clear error (an empty workbook cannot be saved).
    """

    slug = "pdf-to-xlsx"
    input_formats = ("pdf",)
    output_format = "xlsx"

    def convert(self, source_path: Path, destination_dir: Path) -> Path:
        destination_dir.mkdir(parents=True, exist_ok=True)
        output_path = destination_dir / f"{source_path.stem}.xlsx"

        logger.info("pdf_to_xlsx.convert.start", extra={"source": str(source_path)})

        doc = fitz.open(source_path)
        try:
            workbook = Workbook()
            workbook.remove(workbook.active)

            tables_written = 0
            for page_index, page in enumerate(doc):
                tables = page.find_tables().tables
                if not tables:
                    continue

                sheet = workbook.create_sheet(title=f"Page {page_index + 1}")
                row_cursor = 1
                for table in tables:
                    for row in table.extract():
                        for col_index, value in enumerate(row, start=1):
                            sheet.cell(
                                row=row_cursor,
                                column=col_index,
                                value=_normalize_cell_value(value),
                            )
                        row_cursor += 1
                    row_cursor += 1  # blank separator row between tables on the same page
                    tables_written += 1

            if tables_written == 0:
                raise ValueError("No tables were found in this PDF.")

            workbook.save(output_path)
        finally:
            doc.close()

        logger.info(
            "pdf_to_xlsx.convert.done",
            extra={"output": str(output_path), "tables_written": tables_written},
        )
        return output_path

    def verify(self, output_path: Path) -> VerificationResult:
        """The `tables_written == 0` check in `convert()` above already
        guarantees at least one non-empty worksheet before a file is ever
        saved (an empty workbook can't be saved at all), so this is
        deliberately NOT a re-implementation of that same "did we find any
        tables" business rule — it's an independent structural check of the
        actual bytes written to disk: is this a valid ZIP/XLSX container
        that openpyxl can still open, with at least one worksheet in it.
        Catches the on-disk file itself being corrupt or truncated, a
        different failure mode than "no tables found".
        """
        if not output_path.exists():
            return VerificationResult(ok=False, reason="output_missing")
        try:
            workbook = load_workbook(output_path, read_only=True)
        except Exception:
            return VerificationResult(ok=False, reason="invalid_xlsx")
        try:
            if not workbook.sheetnames:
                return VerificationResult(ok=False, reason="no_worksheets")
        finally:
            workbook.close()
        return VerificationResult(ok=True)


register_converter(PdfToXlsxConverter())
